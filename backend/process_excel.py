# import openpyxl
# from qdrant_client import QdrantClient
# from qdrant_client.http.models import Filter, SearchRequest, Payload

# # Connect to your Qdrant instance
# qdrant_client = QdrantClient(
#     url="http://localhost:6333"  # Replace with your Qdrant host if different
# )

# # Define the collection name
# COLLECTION_NAME = "mm_collection"  # Replace with your actual collection name

# # Define a function to search for answers in Qdrant
# def search_answer_in_qdrant(question: str) -> str:
#     try:
#         search_result = qdrant_client.search(
#             collection_name=COLLECTION_NAME,
#             query_vector=qdrant_client.embed_text(question),
#             limit=1,
#             with_payload=True
#         )
#         if search_result:
#             # Assuming the answer is stored in the payload under the key "answer"
#             return search_result[0].payload.get("answer", "No answer found")
#         else:
#             return "No relevant answer found"
#     except Exception as e:
#         print(f"Error searching for answer to '{question}': {e}")
#         return "Error retrieving answer"

# # Process the Excel file to populate answers
# def process_excel(file_path: str):
#     # Open the workbook and select the active worksheet
#     wb = openpyxl.load_workbook(file_path)
#     sheet = wb.active

#     # Verify headers in the first row
#     if sheet.cell(row=1, column=1).value != "Questions" or sheet.cell(row=1, column=2).value != "Answers":
#         print("Invalid file format. Ensure the columns are labeled 'Questions' and 'Answers'.")
#         return

#     # Iterate over rows, starting from the second row
#     for row in range(2, sheet.max_row + 1):
#         question_cell = sheet.cell(row=row, column=1)
#         answer_cell = sheet.cell(row=row, column=2)

#         question = question_cell.value
#         if question and not answer_cell.value:  # Process only if the question exists and answer is empty
#             print(f"Processing question: {question}")
#             answer = search_answer_in_qdrant(question)
#             answer_cell.value = answer  # Write the answer to the Excel sheet

#     # Save changes back to the file
#     wb.save(file_path)
#     print(f"Answers populated and file saved: {file_path}")

# # Main execution
# if __name__ == "__main__":
#     import sys
#     if len(sys.argv) < 2:
#         print("Usage: python process_excel_qdrant.py <path_to_excel_file>")
#     else:
#         process_excel(sys.argv[1])

import openpyxl
import json
import sys
from qdrant_client import QdrantClient
from qdrant_client.http.models import Filter, SearchRequest, Payload
from sentence_transformers import SentenceTransformer  # Import SentenceTransformer

def output_json(data):
    """Helper function to ensure all output is JSON formatted"""
    print(json.dumps(data))
    sys.stdout.flush()

try:
    # Connect to your Qdrant instance
    qdrant_client = QdrantClient(
        url="http://localhost:6333"
    )

    # Define the collection name
    COLLECTION_NAME = "mm_collection"

    # Initialize Sentence Transformer model
    model = SentenceTransformer('all-MiniLM-L6-v2')  # Or any other transformer model

    # Define a function to search for answers in Qdrant
    def search_answer_in_qdrant(question: str) -> str:
        try:
            # Use the model to encode the query text into a vector
            query_vector = model.encode([question])[0]  # Get the embedding of the question
            
            # Perform the search in Qdrant using the query vector
            search_result = qdrant_client.search(
                collection_name=COLLECTION_NAME,
                query_vector=query_vector.tolist(),  # Ensure the vector is a list
                limit=1,
                with_payload=True
            )
            
            # Initialize an empty list to store the documents
            documents = []

            if search_result:
                for hit in search_result:
                    # Check if the payload is a dictionary
                    if isinstance(hit.payload, dict):
                        # Retrieve the 'answer' from the payload, or fallback to the string representation of the payload
                        doc = hit.payload.get('answer', str(hit.payload))
                    else:
                        # If the payload isn't a dictionary, convert it to a string
                        doc = str(hit.payload)
                    
                    # Append the document (answer) to the documents list
                    documents.append(doc)
                
                # Join the documents into a single string (if there are multiple documents) and return
                return "\n".join(documents) if documents else "No answer found"
            else:
                return "No relevant answer found"

        except Exception as e:
            # Instead of printing directly, return the error message
            return f"Error retrieving answer: {str(e)}"


    # Process the Excel file to populate answers
    def process_excel(file_path: str):
        try:
            # Open the workbook and select the active worksheet
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            # Verify headers in the first row
            if sheet.cell(row=1, column=1).value != "Questions" or sheet.cell(row=1, column=2).value != "Answers":
                output_json({
                    "success": False,
                    "error": "Invalid file format. Ensure the columns are labeled 'Questions' and 'Answers'."
                })
                return

            # Iterate over rows, starting from the second row
            processed_questions = 0
            for row in range(2, sheet.max_row + 1):
                question_cell = sheet.cell(row=row, column=1)
                answer_cell = sheet.cell(row=row, column=2)

                question = question_cell.value
                if question and not answer_cell.value:
                    answer = search_answer_in_qdrant(question)
                    answer_cell.value = answer
                    processed_questions += 1

            # Save changes back to the file
            wb.save(file_path)
            
            # Return success response with file path
            output_json({
                "success": True,
                "message": f"Processed {processed_questions} questions successfully",
                "filePath": file_path
            })
            
        except Exception as e:
            output_json({
                "success": False,
                "error": str(e)
            })

    # Main execution
    if __name__ == "__main__":
        if len(sys.argv) < 2:
            output_json({
                "success": False,
                "error": "No file path provided"
            })
        else:
            process_excel(sys.argv[1])

except Exception as e:
    # Catch any top-level exceptions and output as JSON
    output_json({
        "success": False,
        "error": f"Initialization error: {str(e)}"
    })
